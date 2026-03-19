import streamlit as st
import pandas as pd
import io
import copy
import math
import warnings
import re
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.logic_credit import process_credit_report
from utils.logic_XS import process_overdue_sales
from utils.style import display_pretty_report 

warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="Take It Easy - 智能办公助手",
    page_icon="✨",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    html { font-size: 18px !important; }
    :root {
        --deepseek-blue: #4d6bfe;
        --deepseek-dark: #2b4cff;
        --btn-gradient: linear-gradient(90deg, #4d6bfe 0%, #2b4cff 100%);
        --bg-color: #f8f9fa;
        --text-main: #1f1f1f;
        --text-sub: #5f6368;
    }
    .stApp { background-color: var(--bg-color); }
    .header-container { text-align: center; padding: 3rem 0 1rem 0; }
    .main-title {
        font-size: 4.5rem !important; font-weight: 800; letter-spacing: -2px; margin: 0;
        background: linear-gradient(90deg, #4285f4, #9b72cb, #d96570);
        background-size: 200% auto; -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        animation: shine 5s linear infinite;
    }
    @keyframes shine { to { background-position: 200% center; } }
    .sub-title { font-size: 1rem; color: var(--text-sub); letter-spacing: 2px; text-transform: uppercase; margin-top: 0.5rem; }
    .greeting-text { font-size: 2rem; font-weight: 300; color: var(--text-main); text-align: center; margin-bottom: 2rem; }
    div[role="radiogroup"] > label > div:first-child { display: none !important; }
    div[role="radiogroup"] { display: flex; justify-content: center; gap: 15px; width: 100%; margin-bottom: 25px; }
    div[role="radiogroup"] label {
        background: white; border: 1px solid #e0e0e0; border-radius: 12px; padding: 15px;
        text-align: center; box-shadow: 0 4px 10px rgba(0,0,0,0.05); cursor: pointer; flex: 1;
        transition: all 0.3s; min-height: 80px; display: flex; align-items: center; justify-content: center;
        font-weight: 600; color: var(--text-sub);
    }
    div[role="radiogroup"] label[data-checked="true"] {
        border: 2px solid transparent !important;
        background: linear-gradient(white, white) padding-box, var(--btn-gradient) border-box !important;
        color: var(--deepseek-blue) !important; transform: translateY(-4px); box-shadow: 0 8px 20px rgba(77, 107, 254, 0.2);
    }
    .info-box {
        background: #ffffff; border-left: 4px solid var(--deepseek-blue); padding: 20px 25px;
        border-radius: 0 8px 8px 0; margin-bottom: 25px; color: #4a4a4a; font-size: 1rem;
        box-shadow: 0 2px 10px rgba(0,0,0,0.03); text-align: left; line-height: 1.8;
    }
    .info-title { font-weight: 700; color: #1f1f1f; margin-bottom: 8px; display: flex; align-items: center; gap: 8px; }
    [data-testid="stFileUploader"] section {
        border-radius: 12px; background-color: white; border: 2px dashed #dbe0ea; padding: 1.5rem;
    }
    [data-testid="stFileUploader"] section:hover { border-color: var(--deepseek-blue); }
    div.stButton > button {
        width: 100%; height: 60px; border-radius: 12px; font-size: 1.2rem; font-weight: 600;
        background: var(--btn-gradient); color: white; border: none; transition: all 0.3s ease; box-shadow: 0 4px 15px rgba(77, 107, 254, 0.3);
    }
    div.stButton > button:hover { transform: scale(1.02); box-shadow: 0 8px 25px rgba(77, 107, 254, 0.4); color: white; }
    #MainMenu, header, footer { visibility: hidden; }
    [data-testid="stPills"] { display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 15px; }
    [data-testid="stPills"] button {
        border-radius: 20px !important; border: 1px solid #e0e0e0 !important; background: white !important;
        color: #5f6368 !important; padding: 6px 20px !important; font-size: 0.95rem !important;
        transition: all 0.2s ease; min-height: 40px !important; height: auto !important;
    }
    [data-testid="stPills"] button[aria-selected="true"] {
        background: var(--btn-gradient) !important; color: white !important; border: none !important;
        box-shadow: 0 4px 12px rgba(77, 107, 254, 0.3); font-weight: 600 !important;
    }
    [data-testid="stPills"] button:hover { border-color: var(--deepseek-blue) !important; color: var(--deepseek-blue) !important; transform: translateY(-1px); }
    [data-testid="stPills"] button[aria-selected="true"]:hover { color: white !important; transform: translateY(-1px); }           
</style>
""", unsafe_allow_html=True)

def read_excel_safe(file_stream):
    try:
        file_stream.seek(0)
        df = pd.read_excel(file_stream, sheet_name="WSBZJQKB", dtype={'合同编号': str})
        if '合同编号' not in df.columns:
            file_stream.seek(0)
            df_temp = pd.read_excel(file_stream, sheet_name="WSBZJQKB", header=None, nrows=200)
            header_idx = -1
            for idx, row in df_temp.iterrows():
                if "合同编号" in row.values:
                    header_idx = idx
                    break
            if header_idx != -1:
                file_stream.seek(0)
                df = pd.read_excel(file_stream, sheet_name="WSBZJQKB", header=header_idx, dtype={'合同编号': str})
            else:
                raise ValueError("在文件前200行中无法找到包含'合同编号'的标题行，请检查文件格式。")
        return df
    except Exception as e: raise e

def fill_original_sheet_columns(ws_original, df_data):
    try:
        col_reason = get_column_by_name(ws_original, "逾期具体原因")
        col_type = get_column_by_name(ws_original, "逾期原因分类")
        col_client = get_column_by_name(ws_original, "客户")
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        if col_reason and col_type and col_client:
             for i, row_cells in enumerate(ws_original.iter_rows(min_row=2), start=0):
                if i >= len(df_data): break
                cell_reason = row_cells[col_reason - 1]
                cell_type = row_cells[col_type - 1]
                cell_client = row_cells[col_client - 1]
                new_reason_val = df_data.iloc[i].get("逾期具体原因_新", "")
                new_type_val = df_data.iloc[i].get("逾期原因分类_新", "")
                if cell_reason.value is None or str(cell_reason.value).strip() == "":
                    cell_reason.value = new_reason_val
                    if cell_client.has_style:
                        cell_reason.font = copy.copy(cell_client.font)
                        cell_reason.fill = copy.copy(cell_client.fill)
                    cell_reason.alignment = left_align
                    cell_reason.border = thin_border
                if cell_type.value is None or str(cell_type.value).strip() == "":
                    cell_type.value = new_type_val
                    if cell_client.has_style:
                        cell_type.font = copy.copy(cell_client.font)
                        cell_type.fill = copy.copy(cell_client.fill)
                    cell_type.alignment = left_align
                    cell_type.border = thin_border
        for row in ws_original.iter_rows():
            ws_original.row_dimensions[row[0].row].height = 24.5
            for cell in row:
                cell.border = thin_border
                if cell.alignment:
                    new_align = copy.copy(cell.alignment)
                    new_align.vertical = 'center'
                    cell.alignment = new_align
                else: cell.alignment = Alignment(vertical='center')
    except Exception as e: pass

def get_true_column_width(value):
    if value is None: return 0
    str_val = str(value)
    width = 0
    for char in str_val:
        if ord(char) > 255: width += 2.1
        elif char.isupper() or char.isdigit(): width += 1.2
        else: width += 1.0
    return width

def auto_fit_columns(worksheet, min_width=10, max_width=60):
    custom_widths = {"序号": 6, "业务部门": 14, "合同编号": 28, "客户": 35, "品种": 10, "合同数量": 14, "合同单价": 14, "合同金额": 16, "应收保证金日期": 18, "应收保证金比例": 16, "应收保证金金额": 18, "已收定金/预收款": 18, "逾期初始保证金金额": 22}
    for col in worksheet.columns:
        column_letter = get_column_letter(col[0].column)
        header_text = str(col[0].value).strip() if col[0].value else ""
        matched_width = None
        for key, width in custom_widths.items():
            if key in header_text:
                matched_width = width
                break
        if matched_width:
            worksheet.column_dimensions[column_letter].width = matched_width
            continue
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    cell_width = get_true_column_width(cell.value)
                    if cell_width > max_length: max_length = cell_width
            except: pass
        adjusted_width = min(max(max_length + 3, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def find_header_row(worksheet):
    try:
        max_search_rows = min(200, worksheet.max_row)
        critical_field = "合同编号"
        for row_idx in range(1, max_search_rows + 1):
            row_values = []
            for col_idx in range(1, min(20, worksheet.max_column) + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                if cell_value: row_values.append(str(cell_value).strip())
            for val in row_values:
                if critical_field in val: return row_idx
        for row_idx in range(1, max_search_rows + 1):
            for col_idx in range(1, min(20, worksheet.max_column) + 1):
                val = str(worksheet.cell(row_idx, col_idx).value or "")
                if "业务部门" in val: return row_idx
        return 1
    except: return 1

def remove_empty_rows(worksheet):
    try:
        header_row = find_header_row(worksheet)
        if header_row > 1:
            rows_to_delete = header_row - 1
            worksheet.delete_rows(1, rows_to_delete)
            return True
        return True
    except: return False

def get_column_by_name(worksheet, column_names):
    if isinstance(column_names, str): column_names = [column_names]
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=1, column=col).value
        if cell_value:
            for col_name in column_names:
                if col_name in str(cell_value).strip(): return col
    return None

def beautify_sheet_common(ws, title_color="BDD7EE"):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid")
    header_font = Font(color="000000", bold=True, size=11)
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in range(2, ws.max_row + 1):
        row_bg_fill = white_fill if row % 2 == 0 else light_fill
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            current_fill = cell.fill
            is_yellow_cell = False
            if current_fill and current_fill.start_color and current_fill.start_color.rgb:
                if str(current_fill.start_color.rgb).endswith("FFFF00"): is_yellow_cell = True
            if not is_yellow_cell: cell.fill = row_bg_fill
            if not cell.font.color or cell.font.color.rgb == "00000000": cell.font = Font(size=10)
            cell.alignment = center_align
    ws.row_dimensions[1].height = 25
    for row in range(2, ws.max_row + 1): ws.row_dimensions[row].height = 22
    ws.freeze_panes = 'A2'

def clean_and_organize_A_sheet(ws_A):
    try:
        columns_to_delete = ["区域公司", "公司名称", "销售类型", "业务模式", "合同提交日期", "合同签订日期", "合同生效日期", "出库数量", "是否约定保证金条款", "合同约定几个工作日收取", "已收货款金额（不含保证金）", "逾期具体原因", "逾期原因分类", "逾期具体原因_新", "逾期原因分类_新"]
        cols_found = []
        for col in range(1, ws_A.max_column + 1):
            val = str(ws_A.cell(row=1, column=col).value)
            for target in columns_to_delete:
                if target in val:
                    cols_found.append(col)
                    break
        for col_idx in sorted(cols_found, reverse=True): ws_A.delete_cols(col_idx, 1)
        data = list(ws_A.values)
        if not data: return False
        headers = data[0]
        df = pd.DataFrame(data[1:], columns=headers)
        date_col = next((c for c in df.columns if "应收保证金日期" in str(c)), None)
        if date_col:
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime('%Y-%m-%d')
            df = df.sort_values(by=date_col)
        dept_col = next((c for c in df.columns if "业务部门" in str(c)), None)
        if dept_col:
            replacements = ['沿海深圳', '食品原料部', '经营部', '中粮贸易（深圳）有限公司-', '（旧）']
            for r in replacements: df[dept_col] = df[dept_col].astype(str).str.replace(r, '', regex=False)
        ws_A.delete_rows(2, ws_A.max_row)
        for r_idx, row in enumerate(df.values, 2):
            for c_idx, val in enumerate(row, 1): ws_A.cell(row=r_idx, column=c_idx, value=val)
        serial_col = get_column_by_name(ws_A, "序号")
        contract_col = get_column_by_name(ws_A, "合同编号")
        if serial_col and contract_col:
            col_letter = get_column_letter(contract_col)
            for r in range(2, ws_A.max_row + 1): ws_A.cell(row=r, column=serial_col, value=f'=SUBTOTAL(103, ${col_letter}$2:{col_letter}{r})')
        numeric_cols = ["合同数量", "合同单价", "合同金额", "应收保证金金额", "已收定金", "逾期初始保证金"]
        for col_name in numeric_cols:
            col_idx = get_column_by_name(ws_A, col_name)
            if col_idx:
                for r in range(2, ws_A.max_row + 1):
                    cell = ws_A.cell(row=r, column=col_idx)
                    try:
                        if cell.value:
                            cell.value = float(cell.value)
                            cell.number_format = '0.00'
                    except: pass
        pct_col = get_column_by_name(ws_A, "应收保证金比例")
        if pct_col:
            for r in range(2, ws_A.max_row + 1):
                cell = ws_A.cell(row=r, column=pct_col)
                try:
                    if cell.value:
                        cell.value = float(cell.value)
                        cell.number_format = '0%'
                except: pass
        return True
    except: return False

def optimize_A_sheet_formatting(ws_A):
    try:
        today = datetime.now().date()
        date_column = get_column_by_name(ws_A, "应收保证金日期")
        if date_column:
            dark_red_font = Font(color="8B0000")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row in range(2, ws_A.max_row + 1):
                cell = ws_A.cell(row=row, column=date_column)
                try:
                    if cell.value:
                        cell_date_str = str(cell.value)
                        cell_date = None
                        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"]:
                            try:
                                cell_date = datetime.strptime(cell_date_str, fmt).date()
                                break
                            except: continue
                        if cell_date and cell_date <= today:
                            for col in range(1, ws_A.max_column + 1): ws_A.cell(row=row, column=col).font = dark_red_font
                            if cell_date < today: cell.fill = yellow_fill
                except: continue
        beautify_sheet_common(ws_A, title_color="BDD7EE")
        right_align_keywords = ["应收保证金日期", "应收保证金比例", "应收保证金金额", "已收定金/预收款", "逾期初始保证金金额"]
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        for keyword in right_align_keywords:
            col_idx = get_column_by_name(ws_A, keyword)
            if col_idx:
                for row in range(2, ws_A.max_row + 1): ws_A.cell(row=row, column=col_idx).alignment = right_align
        auto_fit_columns(ws_A)
    except: pass

def create_A_summary_sheet(workbook, ws_A, today_date_str):
    try:
        if "A类逾期明细汇总" in workbook.sheetnames: del workbook["A类逾期明细汇总"]
        ws_summary = workbook.create_sheet("A类逾期明细汇总")
        ws_summary.append(["业务部门", "提醒内容"])
        today_date = datetime.strptime(today_date_str, "%Y.%m.%d")
        yesterday_str = (today_date - timedelta(days=1)).strftime("%m月%d日")
        business_dept_col = get_column_by_name(ws_A, "业务部门")
        date_col = get_column_by_name(ws_A, "应收保证金日期")
        if not business_dept_col or not date_col: return False, []
        dept_stats = {}
        for row in range(2, ws_A.max_row + 1):
            dept_name = ws_A.cell(row=row, column=business_dept_col).value
            if not dept_name: dept_name = "未知部门"
            if dept_name not in dept_stats: dept_stats[dept_name] = {'total': 0, 'yellow_cells': 0, 'non_yellow_cells': 0}
            dept_stats[dept_name]['total'] += 1
            cell_fill = ws_A.cell(row=row, column=date_col).fill
            is_yellow = False
            if cell_fill and cell_fill.start_color and cell_fill.start_color.rgb:
                if str(cell_fill.start_color.rgb).endswith("FFFF00"): is_yellow = True
            if is_yellow: dept_stats[dept_name]['yellow_cells'] += 1
            else: dept_stats[dept_name]['non_yellow_cells'] += 1
        logs = []
        row_idx = 2
        for dept_name, stats in dept_stats.items():
            if stats['yellow_cells'] > 0:
                reminder_text = f"【逾期初始保证金】各位领导同事，截至{yesterday_str}，{dept_name}经营部初始保证金{stats['yellow_cells']}笔逾期，{stats['non_yellow_cells']}笔即将到期，请核对并及时催收，谢谢！ @所有人"
            else:
                reminder_text = f"【逾期初始保证金】各位领导同事，截至{yesterday_str}，{dept_name}经营部初始保证金{stats['non_yellow_cells']}笔即将到期，请核对并及时催收，谢谢！ @所有人"
            ws_summary.cell(row=row_idx, column=1, value=dept_name)
            ws_summary.cell(row=row_idx, column=2, value=reminder_text)
            clean_log = reminder_text.replace('\n', '').replace('\r', '')
            logs.append(f"📌 {dept_name}: {clean_log}")
            row_idx += 1
        beautify_sheet_common(ws_summary, title_color="BDD7EE")
        dept_len = 0
        for cell in ws_summary['A']:
            val_len = get_true_column_width(cell.value)
            if val_len > dept_len: dept_len = val_len
        ws_summary.column_dimensions['A'].width = min(max(dept_len + 4, 15), 40)
        fixed_text_width = 90
        ws_summary.column_dimensions['B'].width = fixed_text_width
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for row in range(2, ws_summary.max_row + 1): ws_summary.cell(row=row, column=2).alignment = left_align
        for row in range(2, ws_summary.max_row + 1):
            cell_val = str(ws_summary.cell(row=row, column=2).value or "")
            text_width = get_true_column_width(cell_val)
            estimated_lines = math.ceil(text_width / (fixed_text_width - 5))
            if estimated_lines <= 1: row_height = 25
            else: row_height = 20 + (estimated_lines - 1) * 18
            ws_summary.row_dimensions[row].height = row_height
        return True, logs
    except: return False, []

def process_margin_deposit_logic(current_file, prev_file):
    try:
        book = openpyxl.load_workbook(current_file)
        if "WSBZJQKB" in book.sheetnames: remove_empty_rows(book["WSBZJQKB"])
        temp_stream = io.BytesIO()
        book.save(temp_stream)
        temp_stream.seek(0)
        df_today = read_excel_safe(temp_stream)
        df_last = read_excel_safe(prev_file)
        df_today = df_today.loc[:, ~df_today.columns.str.contains('^Unnamed')]
        df_last = df_last.loc[:, ~df_last.columns.str.contains('^Unnamed')]
        mapping = {}
        for _, row in df_last.iterrows():
            cid = str(row.get('合同编号', '')).strip()
            if cid and cid != 'nan': mapping[cid] = {'r': row.get('逾期具体原因', ''), 'c': row.get('逾期原因分类', '')}
        df_today['合同编号'] = df_today['合同编号'].astype(str).str.strip()
        df_today["逾期具体原因_新"] = df_today["合同编号"].apply(lambda x: mapping.get(x, {}).get('r', ''))
        df_today["逾期原因分类_新"] = df_today["合同编号"].apply(lambda x: mapping.get(x, {}).get('c', ''))
        mask_empty = df_today["逾期原因分类_新"] == ""
        if mask_empty.any():
            clause_col = "是否约定保证金条款"
            if clause_col in df_today.columns:
                df_today.loc[mask_empty & (df_today[clause_col] == "是"), ["逾期具体原因_新", "逾期原因分类_新"]] = ["保证金待收取，已催收", "A实际已逾期：指未按合同约定及时足额支付初始保证金。"]
                df_today.loc[mask_empty & (df_today[clause_col] == "否"), ["逾期具体原因_新", "逾期原因分类_新"]] = ["合同未约定收取保证金", "C无需收取保证金：指政策性业务、对养殖户销售业务、分合同、公司批准免收保证金客户的。此类要写明不收取保证金的具体原因。"]
        temp_stream.seek(0)
        book = openpyxl.load_workbook(temp_stream)
        for s in ["WSBZJQKB_Processed", "A类逾期明细", "A类逾期明细汇总"]:
            if s in book.sheetnames: del book[s]
        ws_proc = book.create_sheet("WSBZJQKB_Processed")
        for r in dataframe_to_rows(df_today, index=False, header=True): ws_proc.append(r)
        df_A = df_today[df_today["逾期原因分类_新"] == "A实际已逾期：指未按合同约定及时足额支付初始保证金。"].copy()
        ws_A = book.create_sheet("A类逾期明细")
        for r in dataframe_to_rows(df_A, index=False, header=True): ws_A.append(r)
        clean_and_organize_A_sheet(ws_A)
        optimize_A_sheet_formatting(ws_A)
        today_str = datetime.now().strftime("%Y.%m.%d")
        success, logs = create_A_summary_sheet(book, ws_A, today_str)
        if "WSBZJQKB" in book.sheetnames: fill_original_sheet_columns(book["WSBZJQKB"], df_today)
        if "WSBZJQKB_Processed" in book.sheetnames: del book["WSBZJQKB_Processed"]
        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        return output, logs
    except Exception as e:
        import traceback
        return None, [f"❌ 处理出错: {str(e)}", traceback.format_exc()]

def smart_format_money_zj(value):
    try:
        if pd.isna(value) or value is None: return "0"
        val_float = float(value)
        if abs(val_float) < 0.000001: return "0"
        val_round = round(val_float)
        if val_round == 0: return f"{val_float:.2f}"
        else: return str(val_round)
    except: return str(value)

def smart_format_volume_zj(value, unit="万吨"):
    try:
        if pd.isna(value) or value is None: return f"0{unit}"
        val_float = float(value)
        if 0.0001 <= abs(val_float) < 0.005:
            val_tons = round(val_float * 10000)
            return f"{val_tons}吨"
        s = f"{val_float:.2f}"
        s = s.rstrip('0').rstrip('.')
        if not s: s = "0"
        return f"{s}{unit}"
    except: return f"{str(value)}{unit}"

def smart_format_date_zj(date_obj):
    try:
        if pd.isna(date_obj): return ""
        if isinstance(date_obj, str):
            try: date_obj = pd.to_datetime(date_obj)
            except: return date_obj
        return f"{date_obj.month}月{date_obj.day}日"
    except: return str(date_obj)

def format_number_with_thousands_zj(value):
    try:
        if pd.isna(value) or value is None: return "0"
        num_value = float(value)
        if num_value == 0: return "0"
        int_value = round(num_value)
        if int_value == 0 and abs(num_value) > 0: return f"{num_value:.2f}"
        return f"{int_value:,}"
    except Exception as e: return str(value)

def find_header_row_zj(ws, max_rows_to_check=30):
    key_columns = ['序号', '大区', '经营部', '品种', '客户名称', '合同编号', '合同数量', '合同单价', '调整后待追加保证金金额', '逾期天数', '调整后待执行数量']
    header_candidates = []
    for row_idx in range(1, min(max_rows_to_check, ws.max_row) + 1):
        row_values = []
        key_count = 0
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).strip()
                row_values.append(cell_str)
                for key in key_columns:
                    if key in cell_str:
                        key_count += 1
                        break
        if key_count >= 2 or len(row_values) > 5:
            non_null_count = sum(1 for v in row_values if v and str(v).strip())
            header_candidates.append({'row': row_idx, 'key_count': key_count, 'non_null_count': non_null_count})
    if not header_candidates: return 5
    header_candidates.sort(key=lambda x: (x['key_count'], x['non_null_count']), reverse=True)
    return header_candidates[0]['row']

def get_column_mapping_zj(ws, header_row):
    column_mapping = {}
    reverse_mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        col_name = cell.value
        if not col_name or str(col_name).strip() == '': col_name = f'Unnamed_{col_idx}'
        else: col_name = str(col_name).strip()
        column_mapping[col_idx] = col_name
        reverse_mapping[col_name] = col_idx
    return column_mapping, reverse_mapping

def copy_cell_style_zj(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)
    return target_cell

def apply_excel_like_filtering_zj(ws_original, ws_processed):
    try:
        header_row_idx = find_header_row_zj(ws_original)
        column_mapping, _ = get_column_mapping_zj(ws_original, header_row_idx)

        for col_idx in range(1, ws_original.max_column + 1):
            source = ws_original.cell(row=header_row_idx, column=col_idx)
            target = ws_processed.cell(row=1, column=col_idx)
            target.value = source.value
            copy_cell_style_zj(source, target)
        
        data_rows = []
        for row_idx in range(header_row_idx + 1, ws_original.max_row + 1):
            row_data = []
            for col_idx in range(1, ws_original.max_column + 1):
                row_data.append(ws_original.cell(row=row_idx, column=col_idx).value)
            data_rows.append((row_idx, row_data))
        
        am_idx = ak_idx = as_idx = b_idx = None
        for idx, name in column_mapping.items():
            s = str(name)
            if '调整后待追加保证金金额' in s: am_idx = idx
            elif '待执行数量调整原因分类' in s: ak_idx = idx
            elif '逾期未回款原因分类' in s: as_idx = idx
            elif '大区' in s and '玉米中心' not in s: b_idx = idx

        if not all([am_idx, ak_idx, as_idx, b_idx]): return [], column_mapping

        filtered_rows = []
        for r_idx, r_data in data_rows:
            include = True
            try:
                val = float(r_data[am_idx-1]) if r_data[am_idx-1] is not None else 0
                if val <= 0.01: include = False
            except: include = False

            if include:
                val = str(r_data[ak_idx-1]) if r_data[ak_idx-1] else ""
                if val in ["合同不再继续执行", "合同约定免收追加保证金"]: include = False
            
            if include:
                val = str(r_data[as_idx-1]) if r_data[as_idx-1] else ""
                if val == "C:追加保证金实际已收到，尚未认领": include = False

            if include:
                val = str(r_data[b_idx-1]) if r_data[b_idx-1] else ""
                if val == "玉米中心": include = False

            if include:
                filtered_rows.append((r_idx, r_data))

        curr_row = 2
        for s_idx, r_data in filtered_rows:
            for c_idx, val in enumerate(r_data, 1):
                s_cell = ws_original.cell(row=s_idx, column=c_idx)
                t_cell = ws_processed.cell(row=curr_row, column=c_idx)
                t_cell.value = val
                copy_cell_style_zj(s_cell, t_cell)
            curr_row += 1

        for col in range(1, ws_original.max_column + 1):
            col_letter = get_column_letter(col)
            if ws_original.column_dimensions[col_letter].width:
                ws_processed.column_dimensions[col_letter].width = ws_original.column_dimensions[col_letter].width
        
        ws_processed.freeze_panes = 'A2'
        return filtered_rows, column_mapping
    except: return [], {}

def generate_analysis_report_zj(df_processed, today_display):
    try:
        d_col = b_col = exec_qty_col = am_col = trigger_date_col = an_col = deposit_type_col = None
        for col_name in df_processed.columns:
            col_str = str(col_name)
            if '细分品种' in col_str: d_col = col_name
            elif '大区' in col_str and '玉米中心' not in col_str: b_col = col_name
            elif '调整后待执行数量' in col_str: exec_qty_col = col_name
            elif '调整后待追加保证金金额' in col_str: am_col = col_name
            elif ('追加保证金触发日期' in col_str or '触发日期' in col_str) and '到期' not in col_str: trigger_date_col = col_name
            elif '逾期' in col_str and '天' in col_str: an_col = col_name
            elif '保证金类型' in col_str: deposit_type_col = col_name

        if not exec_qty_col or not am_col: return "分析报告生成失败：缺少必要的列数据。"

        df_processed[exec_qty_col] = pd.to_numeric(df_processed[exec_qty_col], errors='coerce')
        df_processed[am_col] = pd.to_numeric(df_processed[am_col], errors='coerce')
        if an_col: df_processed[an_col] = pd.to_numeric(df_processed[an_col], errors='coerce')
        if trigger_date_col: df_processed[trigger_date_col] = pd.to_datetime(df_processed[trigger_date_col], errors='coerce')

        total_contracts = len(df_processed)
        total_exec_qty = df_processed[exec_qty_col].sum() / 10000
        total_am_amount = df_processed[am_col].sum()

        deposit_amount_parts = []
        if deposit_type_col:
            df_processed[deposit_type_col] = df_processed[deposit_type_col].astype(str)
            down_deposit = df_processed[df_processed[deposit_type_col].str.contains('跌价', na=False)][am_col].sum()
            up_deposit = df_processed[df_processed[deposit_type_col].str.contains('涨价', na=False)][am_col].sum()
            if down_deposit > 0.000001: deposit_amount_parts.append(f"应收取跌价保证金{smart_format_money_zj(down_deposit)}万元")
            if up_deposit > 0.000001: deposit_amount_parts.append(f"应收取涨价保证金{smart_format_money_zj(up_deposit)}万元")
            if not deposit_amount_parts: deposit_amount_str = f"应收取追加保证金{smart_format_money_zj(total_am_amount)}万元"
            else: deposit_amount_str = "、".join(deposit_amount_parts)
        else: deposit_amount_str = f"应收取追加保证金{smart_format_money_zj(total_am_amount)}万元"

        product_summary = []
        if d_col:
            for product, group in df_processed.groupby(d_col):
                amt = group[am_col].sum()
                if amt > 0.000001: product_summary.append(f"{product}{smart_format_money_zj(amt)}万元")
        product_summary_str = "，".join(product_summary)

        trigger_date_summary = []
        trigger_date_summary_str = ""
        overdue_contracts = 0
        overdue_amount = 0
        if trigger_date_col:
            df_sorted = df_processed.sort_values(by=trigger_date_col)
            if an_col:
                mask = df_processed[an_col] > 0
                overdue_contracts = mask.sum()
                if overdue_contracts > 0: overdue_amount = df_processed.loc[mask, am_col].sum()
            for date, group in df_sorted.groupby(trigger_date_col):
                if pd.notnull(date):
                    date_str = smart_format_date_zj(date)
                    d_amt = group[am_col].sum()
                    o_str = ""
                    if an_col:
                        od = group[an_col].dropna()
                        if not od.empty and od.max() > 0: o_str = f"（逾期{int(od.max())}天）"
                    if d_amt > 0.000001: trigger_date_summary.append(f"{date_str}触发{smart_format_money_zj(d_amt)}万元{o_str}")
            trigger_date_summary_str = "，".join(trigger_date_summary)

        region_summary = []
        if b_col:
            r_data = []
            for region, group in df_processed.groupby(b_col):
                r_data.append({'region': region, 'contracts': len(group), 'exec_qty': group[exec_qty_col].sum()/10000, 'am_amount': group[am_col].sum()})
            r_data.sort(key=lambda x: x['am_amount'], reverse=True)
            for i, r in enumerate(r_data, 1):
                region_summary.append(f"{i}、{r['region']}：{r['contracts']}笔，待执行数量{smart_format_volume_zj(r['exec_qty'])}，需追加保证金金额{smart_format_money_zj(r['am_amount'])}万元。")
        region_summary_str = "\n".join(region_summary)

        report_base = f"""截至{today_display}，存续追加保证金合同{total_contracts}笔，对应待执行量{smart_format_volume_zj(total_exec_qty)}，{deposit_amount_str}"""
        if product_summary_str: report_base += f"。分品种看，{product_summary_str}"
        if overdue_contracts > 0: report_base += f"。其中，{overdue_contracts}笔合同已逾期，逾期金额{smart_format_money_zj(overdue_amount)}万元"
        if trigger_date_summary_str:
            sep = "。" if overdue_contracts > 0 else "。其中，"
            report_base += f"{sep}{trigger_date_summary_str}"
        # ⚠️ 在此处强制另起一行生成“分大区情况如下”
        return report_base + f"。\n\n分大区情况如下：\n{region_summary_str}"
    except: return "分析报告生成失败。"

def generate_customer_analysis_report_zj(df_processed, today_display):
    try:
        c_col = b_col = exec_qty_col = am_col = an_col = deposit_type_col = None
        for col_name in df_processed.columns:
            col_str = str(col_name)
            if '客户' in col_str and '名称' in col_str: c_col = col_name
            elif '大区' in col_str and '玉米中心' not in col_str: b_col = col_name
            elif '调整后待执行数量' in col_str: exec_qty_col = col_name
            elif '调整后待追加保证金金额' in col_str: am_col = col_name
            elif '逾期' in col_str and '天' in col_str: an_col = col_name
            elif '保证金类型' in col_str: deposit_type_col = col_name

        if not c_col or not am_col: return "客户分析报告生成失败：缺少必要的列数据。"

        df_processed[exec_qty_col] = pd.to_numeric(df_processed[exec_qty_col], errors='coerce')
        df_processed[am_col] = pd.to_numeric(df_processed[am_col], errors='coerce')
        if an_col: df_processed[an_col] = pd.to_numeric(df_processed[an_col], errors='coerce')

        total_am_fmt = format_number_with_thousands_zj(df_processed[am_col].sum())
        report_header = f"截至{today_display}，存续追加保证金合同{len(df_processed)}笔，待执行数量{smart_format_volume_zj(df_processed[exec_qty_col].sum()/10000)}，需追加保证金金额{total_am_fmt}万元。"

        c_data = []
        for customer, group in df_processed.groupby(c_col):
            if pd.isna(customer) or customer == "": continue
            regions_str = "、".join([str(r) for r in (group[b_col].dropna().unique() if b_col else []) if pd.notna(r)])
            d_types = ""
            if deposit_type_col:
                dt = group[deposit_type_col].dropna().unique()
                dt_str = "、".join([str(d) for d in dt if pd.notna(d) and str(d).strip() != ""])
                if dt_str: d_types = f"{dt_str}，"
            max_od = group[an_col].max() if an_col else 0
            if pd.isna(max_od): max_od = 0
            
            c_data.append({
                'customer': customer, 'regions': regions_str, 'contracts': len(group),
                'exec_qty': group[exec_qty_col].sum()/10000, 'am_amount': group[am_col].sum(),
                'max_overdue': max_od, 'am_fmt': format_number_with_thousands_zj(group[am_col].sum()),
                'max_od_str': str(round(max_od)), 'd_types': d_types
            })

        c_data.sort(key=lambda x: (-x['max_overdue'], -x['contracts'], -x['exec_qty']))
        c_summary = []
        for i, info in enumerate(c_data, 1):
            od_s = f"，最长逾期{info['max_od_str']}天" if info['max_overdue'] > 0 else ""
            line = f"{i}、{info['regions']}：{info['contracts']}笔，{info['customer']}，{info['d_types']}待执行数量{smart_format_volume_zj(info['exec_qty'])}，需追加保证金金额{info['am_fmt']}万元{od_s}。"
            c_summary.append(line)
        return f"{report_header}\n\n分客户情况如下：\n{'\n'.join(c_summary)}"
    except: return "客户分析报告生成失败。"

def generate_region_department_report_zj(df_region, today_display, region_name):
    try:
        exec_qty_col = am_col = d_col = trigger_date_col = an_col = dept_col = deposit_type_col = None
        for col_name in df_region.columns:
            col_str = str(col_name)
            if '调整后待执行数量' in col_str: exec_qty_col = col_name
            elif '调整后待追加保证金金额' in col_str: am_col = col_name
            elif '细分品种' in col_str: d_col = col_name
            elif ('追加保证金触发日期' in col_str or '触发日期' in col_str) and '到期' not in col_str: trigger_date_col = col_name
            elif '逾期' in col_str and '天' in col_str: an_col = col_name
            elif '经营部' in col_str: dept_col = col_name
            elif '保证金类型' in col_str: deposit_type_col = col_name

        if not exec_qty_col or not am_col: return f"{region_name}大区报告生成失败：缺少必要列数据。"

        df_region[exec_qty_col] = pd.to_numeric(df_region[exec_qty_col], errors='coerce')
        df_region[am_col] = pd.to_numeric(df_region[am_col], errors='coerce')
        if an_col: df_region[an_col] = pd.to_numeric(df_region[an_col], errors='coerce')
        if trigger_date_col: df_region[trigger_date_col] = pd.to_datetime(df_region[trigger_date_col], errors='coerce')

        total_exec_qty = df_region[exec_qty_col].sum() / 10000
        total_am_amount = df_region[am_col].sum()

        deposit_amount_parts = []
        if deposit_type_col:
            df_region[deposit_type_col] = df_region[deposit_type_col].astype(str)
            down_deposit = df_region[df_region[deposit_type_col].str.contains('跌价', na=False)][am_col].sum()
            up_deposit = df_region[df_region[deposit_type_col].str.contains('涨价', na=False)][am_col].sum()
            if down_deposit > 0.000001: deposit_amount_parts.append(f"应收取跌价保证金{smart_format_money_zj(down_deposit)}万元")
            if up_deposit > 0.000001: deposit_amount_parts.append(f"应收取涨价保证金{smart_format_money_zj(up_deposit)}万元")
            if not deposit_amount_parts: deposit_amount_str = f"应收取追加保证金{smart_format_money_zj(total_am_amount)}万元"
            else: deposit_amount_str = "、".join(deposit_amount_parts)
        else: deposit_amount_str = f"应收取追加保证金{smart_format_money_zj(total_am_amount)}万元"

        prod_summary_str = ""
        if d_col:
            prods = []
            for p, g in df_region.groupby(d_col):
                amt = g[am_col].sum()
                if amt > 0.000001: prods.append(f"{p}{smart_format_money_zj(amt)}万元")
            prod_summary_str = "，".join(prods)

        trigger_str = ""
        if trigger_date_col:
            t_sums = []
            df_sorted = df_region.sort_values(by=trigger_date_col)
            for date, group in df_sorted.groupby(trigger_date_col):
                if pd.notnull(date):
                    d_amt = group[am_col].sum()
                    o_str = ""
                    if an_col:
                        od = group[an_col].dropna()
                        if not od.empty and od.max() > 0: o_str = f"（逾期{int(od.max())}天）"
                    if d_amt > 0.000001: t_sums.append(f"{smart_format_date_zj(date)}触发{smart_format_money_zj(d_amt)}万元{o_str}")
            trigger_str = "，".join(t_sums)

        overdue_contracts = 0
        overdue_amount = 0
        if an_col:
            mask = df_region[an_col] > 0
            overdue_contracts = mask.sum()
            if overdue_contracts > 0: overdue_amount = df_region.loc[mask, am_col].sum()

        dept_str = ""
        if dept_col:
            d_data = []
            for dept, group in df_region.groupby(dept_col):
                d_data.append({'dept': dept, 'contracts': len(group), 'exec_qty': group[exec_qty_col].sum()/10000, 'am_amount': group[am_col].sum()})
            d_data.sort(key=lambda x: x['am_amount'], reverse=True)
            d_lines = []
            for i, d in enumerate(d_data, 1):
                name = d['dept'] if pd.notna(d['dept']) and d['dept'] != "" else "未知经营部"
                d_lines.append(f"{i}、{name}：{d['contracts']}笔，待执行数量{smart_format_volume_zj(d['exec_qty'])}，需追加保证金金额{smart_format_money_zj(d['am_amount'])}万元。")
            dept_str = "\n".join(d_lines)

        report_base = f"""截至{today_display}，{region_name}存续追加保证金合同{len(df_region)}笔，对应待执行量{smart_format_volume_zj(total_exec_qty)}，{deposit_amount_str}"""
        if prod_summary_str: report_base += f"。分品种看，{prod_summary_str}"
        if overdue_contracts > 0: report_base += f"。其中，{overdue_contracts}笔合同已逾期，逾期金额{smart_format_money_zj(overdue_amount)}万元"
        if trigger_str:
            sep = "。" if overdue_contracts > 0 else "。其中，"
            report_base += f"{sep}{trigger_str}"
        # ⚠️ 在此处强制另起一行生成“分经营部情况如下”
        return report_base + f"。\n\n分经营部情况如下：\n{dept_str}"
    except: return f"{region_name}大区报告生成失败。"

def generate_region_customer_report_zj(df_region, today_display, region_name):
    try:
        c_col = exec_qty_col = am_col = an_col = dept_col = deposit_type_col = None
        for col_name in df_region.columns:
            col_str = str(col_name)
            if '客户' in col_str and '名称' in col_str: c_col = col_name
            elif '调整后待执行数量' in col_str: exec_qty_col = col_name
            elif '调整后待追加保证金金额' in col_str: am_col = col_name
            elif '逾期' in col_str and '天' in col_str: an_col = col_name
            elif '经营部' in col_str: dept_col = col_name
            elif '保证金类型' in col_str: deposit_type_col = col_name

        if not c_col or not am_col: return f"{region_name}大区客户分析报告生成失败。"

        df_region[exec_qty_col] = pd.to_numeric(df_region[exec_qty_col], errors='coerce')
        df_region[am_col] = pd.to_numeric(df_region[am_col], errors='coerce')
        if an_col: df_region[an_col] = pd.to_numeric(df_region[an_col], errors='coerce')

        total_am_fmt = format_number_with_thousands_zj(df_region[am_col].sum())
        report_header = f"截至{today_display}，{region_name}存续追加保证金合同{len(df_region)}笔，待执行数量{smart_format_volume_zj(df_region[exec_qty_col].sum()/10000)}，需追加保证金金额{total_am_fmt}万元。"

        c_data = []
        for customer, group in df_region.groupby(c_col):
            if pd.isna(customer) or customer == "": continue
            depts_str = ""
            if dept_col:
                depts = group[dept_col].dropna().unique()
                depts_str = "、".join([str(d) for d in depts if pd.notna(d) and str(d).strip() != ""])
            d_types_str = ""
            if deposit_type_col:
                dt = group[deposit_type_col].dropna().unique()
                t_str = "、".join([str(t) for t in dt if pd.notna(t) and str(t).strip() != ""])
                if t_str: d_types_str = f"{t_str}，"
            max_od = group[an_col].max() if an_col else 0
            if pd.isna(max_od): max_od = 0
            c_data.append({
                'customer': customer, 'depts': depts_str, 'contracts': len(group),
                'exec_qty': group[exec_qty_col].sum()/10000, 'am_amount': group[am_col].sum(),
                'max_overdue': max_od, 'am_fmt': format_number_with_thousands_zj(group[am_col].sum()),
                'max_od_str': str(round(max_od)), 'd_types': d_types_str
            })

        c_data.sort(key=lambda x: (-x['max_overdue'], -x['contracts'], -x['exec_qty']))
        lines = []
        for i, info in enumerate(c_data, 1):
            od_s = f"，最长逾期{info['max_od_str']}天" if info['max_overdue'] > 0 else ""
            prefix = f"{i}、{info['depts']}：" if info['depts'] else f"{i}、"
            lines.append(f"{prefix}{info['contracts']}笔，{info['customer']}，{info['d_types']}待执行数量{smart_format_volume_zj(info['exec_qty'])}，需追加保证金金额{info['am_fmt']}万元{od_s}。")
        return f"{report_header}\n\n分客户情况如下：\n{'\n'.join(lines)}"
    except: return f"{region_name}大区客户分析报告生成失败。"

def process_additional_margin_logic(uploaded_file, region_filter):
    logs = []
    try:
        today_display = f"{datetime.now().month}月{datetime.now().day}日"
        
        book = openpyxl.load_workbook(uploaded_file)
        ws_original = book.worksheets[0] 
        
        if '追保处理' in book.sheetnames: del book['追保处理']
        ws_processed = book.create_sheet('追保处理')
        filtered_rows, column_names = apply_excel_like_filtering_zj(ws_original, ws_processed)
        
        if not filtered_rows:
            return None, ["⚠️ 警告：筛选后没有数据行！"], "", ""

        data_for_analysis = []
        for _, row_data in filtered_rows:
            row_dict = {}
            for col_idx, value in enumerate(row_data, 1):
                if col_idx in column_names:
                    row_dict[column_names[col_idx]] = value
            data_for_analysis.append(row_dict)
        df_processed = pd.DataFrame(data_for_analysis)
        
        if '分析报告' in book.sheetnames: del book['分析报告']
        ws_report = book.create_sheet('分析报告')
        
        b_col = next((c for c in df_processed.columns if '大区' in str(c) and '玉米中心' not in str(c)), None)
        report_A = ""
        report_B = ""
        
        if region_filter == "中粮贸易":
            report_A = generate_analysis_report_zj(df_processed, today_display)
            report_B = generate_customer_analysis_report_zj(df_processed, today_display)
        else:
            if not b_col:
                return None, ["❌ 数据中找不到“大区”列，无法进行大区筛选。"], "", ""
            df_region = df_processed[df_processed[b_col] == region_filter].copy()
            if len(df_region) == 0:
                return None, [f"⚠️ 筛选结果中没有包含【{region_filter}】的数据。"], "", ""
            
            report_A = generate_region_department_report_zj(df_region, today_display, region_filter)
            report_B = generate_region_customer_report_zj(df_region, today_display, region_filter)

        ws_report.cell(row=1, column=1, value=report_A)
        ws_report.cell(row=1, column=2, value=report_B)
        
        ws_report.column_dimensions['A'].width = 100
        ws_report.column_dimensions['B'].width = 100
        for row in ws_report.iter_rows():
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.font = Font(size=10, name='宋体')
                    ws_report.row_dimensions[cell.row].height = 200
        ws_report.freeze_panes = 'A2'

        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        
        logs.append(f"✅ 【{region_filter}】分析报告生成成功！")
        return output, logs, report_A, report_B
    except Exception as e:
        import traceback
        return None, [f"❌ 处理出错: {str(e)}", traceback.format_exc()], "", ""

def format_html_content_for_credit(text):
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    list_html = ""
    for line in lines:
        if "情况如下：" in line or "【" in line:
             list_html += f"<div style='font-weight: bold; margin-top: 8px; margin-bottom: 4px; color: #1f1f1f;'>{line.replace('**', '')}</div>"
        else:
             list_html += f"<div style='margin-left: 10px; margin-bottom: 4px; color: #333; line-height: 1.6;'>• {line}</div>"
    return list_html
def main():
    st.markdown("""
        <div class="header-container">
            <h1 class="main-title">Take It Easy</h1>
            <div class="sub-title">Crafted by Xuyingzhe</div>
        </div>
    """, unsafe_allow_html=True)

    col_l, col_center, col_r = st.columns([1, 6, 1])

    with col_center:
        st.markdown('<div class="greeting-text">您好，有什么可以帮到你？</div>', unsafe_allow_html=True)

        function_map = {
            "📈 初始保证金处理": "init_margin",
            "📉 追加保证金处理": "add_margin",
            "⏱️ 逾期销售处理": "overdue_sales",
            "📊 信用风险管理日报": "credit_report"
        }

        mode = st.radio("选择功能", list(function_map.keys()), horizontal=True, label_visibility="collapsed")
        
        # --- 模块 1: 初始保证金处理 ---
        if mode == "📈 初始保证金处理":
            st.markdown("""
            <div class="info-box">
                <div class="info-title">⚠️ 注意事项</div>
                <div style="margin-left: 2px;">
                    <div>请务必同时上传两个文件以便进行数据比对</div>
                    <div style="margin-top: 4px;">原始表单 Sheet 名称必须包含 WSBZJQKB</div>
                    <div style="margin-top: 4px;">生成结果将包含清洗后的明细表及 A 类逾期汇总</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            with c1:
                current_file = st.file_uploader("📂 1. 上传【今日】报表", type=['xlsx'])
            with c2:
                prev_file = st.file_uploader("📂 2. 上传【对照日】报表", type=['xlsx'])
            
            if st.button("🚀 开始处理 / Analyze"):
                if current_file and prev_file:
                    with st.spinner("🤖 正在进行数据比对与清洗，请稍候..."):
                        excel_data, report_logs = process_margin_deposit_logic(current_file, prev_file)
                        
                        if excel_data:
                            st.success("✅ 处理完成！")
                            st.markdown("### 📢 生成的通报文案")
                            for log in report_logs:
                                st.info(log)
                                
                            st.download_button(
                                label=f"📥 下载处理后的报表 ({current_file.name})",
                                data=excel_data,
                                file_name=current_file.name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.error("处理失败，请查看下方错误日志")
                            st.code(report_logs[-1])
                else:
                    st.warning("⚠️ 请确保两个文件都已上传！")
        
        # --- 模块 2: 追加保证金处理 ---
        elif mode == "📉 追加保证金处理":
            st.markdown("""
            <div class="info-box">
                <div class="info-title">⚠️ 注意事项</div>
                <div style="margin-left: 2px;">
                    <div>请务必上传“追加保证金填报表”</div>
                    <div style="margin-top: 4px;">系统将自动进行筛选、数据清洗与报告生成</div>
                    <div style="margin-top: 4px;">下方选择相应大区，即可生成专属定制报告</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown('<div style="margin-bottom: 8px; font-weight: 600; color: #333;">选择报告生成范围</div>', unsafe_allow_html=True)
            region_options = ["中粮贸易", "沿海大区", "沿江大区", "内陆大区", "东北大区"]
            
            selection = st.pills("选择报告生成范围", region_options, default="中粮贸易", label_visibility="collapsed")
            selected_region = selection if selection is not None else "中粮贸易"

            uploaded_file = st.file_uploader("📂 上传【追加保证金填报表】", type=['xlsx'])

            if st.button("🚀 生成报告 / Generate Report"):
                if uploaded_file:
                    with st.spinner(f"🤖 正在为【{selected_region}】生成专属报告..."):
                        output_file, logs, report_a, report_b = process_additional_margin_logic(uploaded_file, selected_region)
                        
                        if output_file:
                            st.success(f"✅ {selected_region}报告生成完成！")
                            
                            # ⚠️ 下载按钮移到了最上方
                            today_mmdd = datetime.now().strftime('%m%d')
                            file_prefix = "" if selected_region == "中粮贸易" else f"{selected_region}"
                            dl_filename = f"{file_prefix}追加保证金填报表{today_mmdd}.xlsx"
                            st.download_button(
                                label=f"📥 下载定制报告 ({dl_filename})",
                                data=output_file,
                                file_name=dl_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            c_a, c_b = st.columns(2)
                            with c_a:
                                # 首段无需加粗
                                display_pretty_report(f"业务单位报告 ({selected_region})", report_a, "#eef5ff", bold_first_para=False)
                            with c_b:
                                display_pretty_report(f"分客户报告 ({selected_region})", report_b, "#fff8e6", bold_first_para=False)
                        else:
                            st.error("处理失败")
                            for l in logs: st.write(l)
                else:
                    st.warning("⚠️ 请先上传文件！")

        # --- 模块 3: 逾期销售处理 ---
        elif mode == "⏱️ 逾期销售处理":
            st.markdown("""
            <div class="info-box">
                <div class="info-title">⚠️ 注意事项</div>
                <div style="margin-left: 2px;">
                    <div>请分别上传【逾期销售（分批次）】和【逾期销售（一次性）】的表格数据</div>
                    <div style="margin-top: 4px;">系统将自动整合数据、计算逾期金额、匹配客户信息</div>
                    <div style="margin-top: 4px;">勾选复选框，可同时生成周报 Word 文档及催收提醒文本</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            need_report = st.checkbox("📝 需要生成【逾期销售周报】(Word格式)", value=False)
            
            col1, col2 = st.columns(2)
            with col1:
                batch_files = st.file_uploader("📂 逾期销售（分批次） [最多6个]", type=["xlsx", "xls"], accept_multiple_files=True, key="batch_upload")
                if batch_files and len(batch_files) > 6:
                    st.warning("⚠️ 分批次文件最多只能上传6个，超出的部分将被忽略。")
                    batch_files = batch_files[:6]
                    
            with col2:
                once_files = st.file_uploader("📂 逾期销售（一次性） [最多6个]", type=["xlsx", "xls"], accept_multiple_files=True, key="once_upload")
                if once_files and len(once_files) > 6:
                    st.warning("⚠️ 一次性文件最多只能上传6个，超出的部分将被忽略。")
                    once_files = once_files[:6]
                    
            if st.button("🚀 开始处理逾期数据", key="btn_xs"):
                if not batch_files and not once_files:
                    st.warning("⚠️ 请至少在一个文件栏中上传数据文件！")
                else:
                    with st.spinner("🤖 正在高速运算并生成报告中..."):
                        excel_io, word_io, collection_text, logs = process_overdue_sales(batch_files, once_files, need_report)
                        
                        if excel_io:
                            st.success("✅ 逾期数据处理成功！")
                            
                            with st.expander("查看处理日志", expanded=False):
                                for log in logs:
                                    st.write(log)
                                    
                            # ⚠️ 下载按钮整体上移至此处
                            st.markdown("### 📥 下载结果文件")
                            dl_col1, dl_col2 = st.columns(2)
                            
                            mmdd_str = datetime.now().strftime('%m%d')
                            yyyymmdd_str = datetime.now().strftime('%Y%m%d')
                            
                            with dl_col1:
                                st.download_button(
                                    label="📊 下载【逾期销售监控表】 (Excel)",
                                    data=excel_io,
                                    file_name=f"逾期销售监控表_{mmdd_str}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                                
                            if need_report and word_io:
                                with dl_col2:
                                    st.download_button(
                                        label="📝 下载【逾期销售周报】 (Word)",
                                        data=word_io,
                                        file_name=f"逾期销售周报_{yyyymmdd_str}.docx",
                                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                        use_container_width=True
                                    )

                            # 预览文本位于最下方，且通过 bold_first_para=True 强制首段加粗
                            if collection_text:
                                st.markdown("### 📢 生成的通报文案")
                                display_pretty_report("💬 催收提醒预览", collection_text, bg_color="#f8f9fa", bold_first_para=True)
                        else:
                            st.error("❌ 处理失败，请检查文件格式是否符合要求。")

        # --- 模块 4: 信用风险管理日报 ---
        elif mode == "📊 信用风险管理日报":
            st.markdown("""
            <div class="info-box">
                <div class="info-title">⚠️ 注意事项</div>
                <div style="margin-left: 2px;">
                    <div>请上传包含「信用风险管理日报」及相应通报 Sheet 的 Excel 文件</div>
                    <div style="margin-top: 4px;">系统将自动抓取逾期数据生成 Word 简报，并导出相关 Sheet</div>
                    <div style="margin-top: 4px;">由于跨平台特性，云端部署时 PDF 导出将降级为高清图片输出</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            uploaded_file = st.file_uploader("📂 上传【信用风险管理日报】Excel 表", type=['xlsx'])
            
            if st.button("🚀 生成报告与导出文件 / Generate"):
                if uploaded_file:
                    with st.spinner("🤖 正在解析 Excel 数据并渲染跨平台文件，请稍候..."):
                        word_bytes, word_text_dict, export_files, logs, env_msg = process_credit_report(uploaded_file)
                        
                        st.info(f"💡 {env_msg}")
                        
                        if word_bytes or export_files:
                            st.success("✅ 任务处理完成！")
                    
                            if word_text_dict:
                                st.markdown("<h3 style='margin-top: 10px; margin-bottom: 20px; color: #1f1f1f;'>信用风险管理日报</h3>", unsafe_allow_html=True)
                                
                                center_themes = {
                                    "玉米": {"bg": "#eef5ff", "bd": "#d1e3ff", "bar": "#4d6bfe"},
                                    "粮谷": {"bg": "#ebf9f1", "bd": "#c3e8d1", "bar": "#28a745"},
                                    "大豆": {"bg": "#fff6e5", "bd": "#ffe2b3", "bar": "#fd7e14"} 
                                }
                                
                                for center_name, content in word_text_dict.items():
                                    theme = center_themes.get(center_name, {"bg": "#fcf8f2", "bd": "#f0e6d2", "bar": "#6c757d"})
                                    html_content = format_html_content_for_credit(content)
                                    
                                    st.markdown(f"""
                                    <div style="background-color: {theme['bg']}; padding: 20px 25px; border-radius: 0 8px 8px 0; border: 1px solid {theme['bd']}; border-left: 4px solid {theme['bar']}; margin-bottom: 20px; box-shadow: 0 2px 10px rgba(0,0,0,0.03);">
                                        {html_content}
                                    </div>
                                    """, unsafe_allow_html=True)
                            
                            st.markdown("### 📥 下载生成文件")
                            dl_cols = st.columns(1 + len(export_files))
                            
                            with dl_cols[0]:
                                if word_bytes:
                                    original_base = os.path.splitext(uploaded_file.name)[0]
                                    st.download_button(
                                        label="📄 下载 Word 报告",
                                        data=word_bytes,
                                        file_name=f"{original_base}.docx",
                                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                        use_container_width=True
                                    )
                                    
                            for i, export_file in enumerate(export_files, 1):
                                with dl_cols[i]:
                                    label = "📉 下载高清图" if export_file["type"] == "png" else "📊 下载 PDF"
                                    mime = "image/png" if export_file["type"] == "png" else "application/pdf"
                                    st.download_button(
                                        label=f"{label} ({export_file['name']})",
                                        data=export_file["data"],
                                        file_name=export_file["name"],
                                        mime=mime,
                                        use_container_width=True
                                    )
                        
                            png_files = [f for f in export_files if f["type"] == "png"]
                            if png_files:
                                st.markdown("#### 👁️ 图片预览")
                                for p_f in png_files:
                                    st.image(p_f["data"], caption=p_f["name"], use_container_width=True)

                        else:
                            st.error("处理失败，未能提取到有效数据。")
                else:
                    st.warning("⚠️ 请先上传 Excel 文件！")

    st.markdown("<div style='text-align:center; color:#ccc; margin-top:50px;'>© 2026 TakeItEasy Tool</div>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
