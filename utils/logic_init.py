import pandas as pd
import io
import copy
import math
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# PART 1: 初始保证金处理逻辑 (XSchushi.txt / app.py 原有逻辑)
# ============================================================================

def read_excel_safe(file_stream):
    try:
        file_stream.seek(0)
        df = pd.read_excel(file_stream, sheet_name="WSBZJQKB", dtype={'合同编号': str})
        if '合同编号' not in df.columns:
            file_stream.seek(0)
            df_temp = pd.read_excel(file_stream, sheet_name="WSBZJQKB", header=None, nrows=200)
            header_idx = -1
            for idx, row in df_temp.iterrows():
                if "合同编号" in row.values:
                    header_idx = idx
                    break
            if header_idx != -1:
                file_stream.seek(0)
                df = pd.read_excel(file_stream, sheet_name="WSBZJQKB", header=header_idx, dtype={'合同编号': str})
            else:
                raise ValueError("在文件前200行中无法找到包含'合同编号'的标题行，请检查文件格式。")
        return df
    except Exception as e:
        raise e

def fill_original_sheet_columns(ws_original, df_data):
    try:
        col_reason = get_column_by_name(ws_original, "逾期具体原因")
        col_type = get_column_by_name(ws_original, "逾期原因分类")
        col_client = get_column_by_name(ws_original, "客户")
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        if col_reason and col_type and col_client:
            for i, row_cells in enumerate(ws_original.iter_rows(min_row=2), start=0):
                if i >= len(df_data): break
                cell_reason = row_cells[col_reason - 1]
                cell_type = row_cells[col_type - 1]
                cell_client = row_cells[col_client - 1]
                new_reason_val = df_data.iloc[i].get("逾期具体原因_新", "")
                new_type_val = df_data.iloc[i].get("逾期原因分类_新", "")
                if cell_reason.value is None or str(cell_reason.value).strip() == "":
                    cell_reason.value = new_reason_val
                    if cell_client.has_style:
                        cell_reason.font = copy.copy(cell_client.font)
                        cell_reason.fill = copy.copy(cell_client.fill)
                    cell_reason.alignment = left_align
                    cell_reason.border = thin_border
                if cell_type.value is None or str(cell_type.value).strip() == "":
                    cell_type.value = new_type_val
                    if cell_client.has_style:
                        cell_type.font = copy.copy(cell_client.font)
                        cell_type.fill = copy.copy(cell_client.fill)
                    cell_type.alignment = left_align
                    cell_type.border = thin_border
        for row in ws_original.iter_rows():
            ws_original.row_dimensions[row[0].row].height = 24.5
            for cell in row:
                cell.border = thin_border
                if cell.alignment:
                    new_align = copy.copy(cell.alignment)
                    new_align.vertical = 'center'
                    cell.alignment = new_align
                else:
                    cell.alignment = Alignment(vertical='center')
    except Exception as e: pass

def get_true_column_width(value):
    if value is None: return 0
    str_val = str(value)
    width = 0
    for char in str_val:
        if ord(char) > 255: width += 2.1
        elif char.isupper() or char.isdigit(): width += 1.2
        else: width += 1.0
    return width

def auto_fit_columns(worksheet, min_width=10, max_width=60):
    custom_widths = {
        "序号": 6, "业务部门": 14, "合同编号": 28, "客户": 35, "品种": 10,
        "合同数量": 14, "合同单价": 14, "合同金额": 16, "应收保证金日期": 18,
        "应收保证金比例": 16, "应收保证金金额": 18, "已收定金/预收款": 18,
        "逾期初始保证金金额": 22
    }
    for col in worksheet.columns:
        column_letter = get_column_letter(col[0].column)
        header_text = str(col[0].value).strip() if col[0].value else ""
        matched_width = None
        for key, width in custom_widths.items():
            if key in header_text:
                matched_width = width
                break
        if matched_width:
            worksheet.column_dimensions[column_letter].width = matched_width
            continue
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    cell_width = get_true_column_width(cell.value)
                    if cell_width > max_length: max_length = cell_width
            except: pass
        adjusted_width = min(max(max_length + 3, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def find_header_row(worksheet):
    try:
        max_search_rows = min(200, worksheet.max_row)
        critical_field = "合同编号"
        for row_idx in range(1, max_search_rows + 1):
            row_values = []
            for col_idx in range(1, min(20, worksheet.max_column) + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                if cell_value: row_values.append(str(cell_value).strip())
            for val in row_values:
                if critical_field in val: return row_idx
        for row_idx in range(1, max_search_rows + 1):
            for col_idx in range(1, min(20, worksheet.max_column) + 1):
                val = str(worksheet.cell(row_idx, col_idx).value or "")
                if "业务部门" in val: return row_idx
        return 1
    except: return 1

def remove_empty_rows(worksheet):
    try:
        header_row = find_header_row(worksheet)
        if header_row > 1:
            rows_to_delete = header_row - 1
            worksheet.delete_rows(1, rows_to_delete)
            return True
        return True
    except: return False

def get_column_by_name(worksheet, column_names):
    if isinstance(column_names, str): column_names = [column_names]
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=1, column=col).value
        if cell_value:
            for col_name in column_names:
                if col_name in str(cell_value).strip(): return col
    return None

def beautify_sheet_common(ws, title_color="BDD7EE"):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid")
    header_font = Font(color="000000", bold=True, size=11)
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in range(2, ws.max_row + 1):
        row_bg_fill = white_fill if row % 2 == 0 else light_fill
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            current_fill = cell.fill
            is_yellow_cell = False
            if current_fill and current_fill.start_color and current_fill.start_color.rgb:
                if str(current_fill.start_color.rgb).endswith("FFFF00"): is_yellow_cell = True
            if not is_yellow_cell: cell.fill = row_bg_fill
            if not cell.font.color or cell.font.color.rgb == "00000000": cell.font = Font(size=10)
            cell.alignment = center_align
    ws.row_dimensions[1].height = 25
    for row in range(2, ws.max_row + 1): ws.row_dimensions[row].height = 22
    ws.freeze_panes = 'A2'

def clean_and_organize_A_sheet(ws_A):
    try:
        columns_to_delete = ["区域公司", "公司名称", "销售类型", "业务模式", "合同提交日期", "合同签订日期", "合同生效日期", "出库数量", "是否约定保证金条款", "合同约定几个工作日收取", "已收货款金额（不含保证金）", "逾期具体原因", "逾期原因分类", "逾期具体原因_新", "逾期原因分类_新"]
        cols_found = []
        for col in range(1, ws_A.max_column + 1):
            val = str(ws_A.cell(row=1, column=col).value)
            for target in columns_to_delete:
                if target in val:
                    cols_found.append(col)
                    break
        for col_idx in sorted(cols_found, reverse=True): ws_A.delete_cols(col_idx, 1)
        data = list(ws_A.values)
        if not data: return False
        headers = data[0]
        df = pd.DataFrame(data[1:], columns=headers)
        date_col = next((c for c in df.columns if "应收保证金日期" in str(c)), None)
        if date_col:
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.strftime('%Y-%m-%d')
            df = df.sort_values(by=date_col)
        dept_col = next((c for c in df.columns if "业务部门" in str(c)), None)
        if dept_col:
            replacements = ['沿海深圳', '食品原料部', '经营部', '中粮贸易（深圳）有限公司-', '（旧）']
            for r in replacements: df[dept_col] = df[dept_col].astype(str).str.replace(r, '', regex=False)
        ws_A.delete_rows(2, ws_A.max_row)
        for r_idx, row in enumerate(df.values, 2):
            for c_idx, val in enumerate(row, 1): ws_A.cell(row=r_idx, column=c_idx, value=val)
        serial_col = get_column_by_name(ws_A, "序号")
        contract_col = get_column_by_name(ws_A, "合同编号")
        if serial_col and contract_col:
            col_letter = get_column_letter(contract_col)
            for r in range(2, ws_A.max_row + 1): ws_A.cell(row=r, column=serial_col, value=f'=SUBTOTAL(103, ${col_letter}$2:{col_letter}{r})')
        numeric_cols = ["合同数量", "合同单价", "合同金额", "应收保证金金额", "已收定金", "逾期初始保证金"]
        for col_name in numeric_cols:
            col_idx = get_column_by_name(ws_A, col_name)
            if col_idx:
                for r in range(2, ws_A.max_row + 1):
                    cell = ws_A.cell(row=r, column=col_idx)
                    try:
                        if cell.value:
                            cell.value = float(cell.value)
                            cell.number_format = '0.00'
                    except: pass
        pct_col = get_column_by_name(ws_A, "应收保证金比例")
        if pct_col:
            for r in range(2, ws_A.max_row + 1):
                cell = ws_A.cell(row=r, column=pct_col)
                try:
                    if cell.value:
                        cell.value = float(cell.value)
                        cell.number_format = '0%'
                except: pass
        return True
    except: return False

def optimize_A_sheet_formatting(ws_A):
    try:
        today = datetime.now().date()
        date_column = get_column_by_name(ws_A, "应收保证金日期")
        if date_column:
            dark_red_font = Font(color="8B0000")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row in range(2, ws_A.max_row + 1):
                cell = ws_A.cell(row=row, column=date_column)
                try:
                    if cell.value:
                        cell_date_str = str(cell.value)
                        cell_date = None
                        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"]:
                            try:
                                cell_date = datetime.strptime(cell_date_str, fmt).date()
                                break
                            except: continue
                        if cell_date and cell_date <= today:
                            for col in range(1, ws_A.max_column + 1): ws_A.cell(row=row, column=col).font = dark_red_font
                            if cell_date < today: cell.fill = yellow_fill
                except: continue
        beautify_sheet_common(ws_A, title_color="BDD7EE")
        right_align_keywords = ["应收保证金日期", "应收保证金比例", "应收保证金金额", "已收定金/预收款", "逾期初始保证金金额"]
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        for keyword in right_align_keywords:
            col_idx = get_column_by_name(ws_A, keyword)
            if col_idx:
                for row in range(2, ws_A.max_row + 1): ws_A.cell(row=row, column=col_idx).alignment = right_align
        auto_fit_columns(ws_A)
    except: pass

def create_A_summary_sheet(workbook, ws_A, today_date_str):
    try:
        if "A类逾期明细汇总" in workbook.sheetnames: del workbook["A类逾期明细汇总"]
        ws_summary = workbook.create_sheet("A类逾期明细汇总")
        ws_summary.append(["业务部门", "提醒内容"])
        today_date = datetime.strptime(today_date_str, "%Y.%m.%d")
        yesterday_str = (today_date - timedelta(days=1)).strftime("%m月%d日")
        business_dept_col = get_column_by_name(ws_A, "业务部门")
        date_col = get_column_by_name(ws_A, "应收保证金日期")
        if not business_dept_col or not date_col: return False, []
        dept_stats = {}
        for row in range(2, ws_A.max_row + 1):
            dept_name = ws_A.cell(row=row, column=business_dept_col).value
            if not dept_name: dept_name = "未知部门"
            if dept_name not in dept_stats: dept_stats[dept_name] = {'total': 0, 'yellow_cells': 0, 'non_yellow_cells': 0}
            dept_stats[dept_name]['total'] += 1
            cell_fill = ws_A.cell(row=row, column=date_col).fill
            is_yellow = False
            if cell_fill and cell_fill.start_color and cell_fill.start_color.rgb:
                if str(cell_fill.start_color.rgb).endswith("FFFF00"): is_yellow = True
            if is_yellow: dept_stats[dept_name]['yellow_cells'] += 1
            else: dept_stats[dept_name]['non_yellow_cells'] += 1
        logs = []
        row_idx = 2
        for dept_name, stats in dept_stats.items():
            if stats['yellow_cells'] > 0:
                reminder_text = f"【逾期初始保证金】各位领导同事，截至{yesterday_str}，{dept_name}经营部初始保证金{stats['yellow_cells']}笔逾期，{stats['non_yellow_cells']}笔即将到期，请核对并及时催收，谢谢！ @所有人"
            else:
                reminder_text = f"【逾期初始保证金】各位领导同事，截至{yesterday_str}，{dept_name}经营部初始保证金{stats['non_yellow_cells']}笔即将到期，请核对并及时催收，谢谢！ @所有人"
            ws_summary.cell(row=row_idx, column=1, value=dept_name)
            ws_summary.cell(row=row_idx, column=2, value=reminder_text)
            clean_log = reminder_text.replace('\n', '').replace('\r', '')
            logs.append(f"📌 {dept_name}: {clean_log}")
            row_idx += 1
        beautify_sheet_common(ws_summary, title_color="BDD7EE")
        dept_len = 0
        for cell in ws_summary['A']:
            val_len = get_true_column_width(cell.value)
            if val_len > dept_len: dept_len = val_len
        ws_summary.column_dimensions['A'].width = min(max(dept_len + 4, 15), 40)
        fixed_text_width = 90
        ws_summary.column_dimensions['B'].width = fixed_text_width
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for row in range(2, ws_summary.max_row + 1): ws_summary.cell(row=row, column=2).alignment = left_align
        for row in range(2, ws_summary.max_row + 1):
            cell_val = str(ws_summary.cell(row=row, column=2).value or "")
            text_width = get_true_column_width(cell_val)
            estimated_lines = math.ceil(text_width / (fixed_text_width - 5))
            if estimated_lines <= 1: row_height = 25
            else: row_height = 20 + (estimated_lines - 1) * 18
            ws_summary.row_dimensions[row].height = row_height
        return True, logs
    except: return False, []

def process_margin_deposit_logic(current_file, prev_file):
    try:
        book = openpyxl.load_workbook(current_file)
        if "WSBZJQKB" in book.sheetnames: remove_empty_rows(book["WSBZJQKB"])
        temp_stream = io.BytesIO()
        book.save(temp_stream)
        temp_stream.seek(0)
        df_today = read_excel_safe(temp_stream)
        df_last = read_excel_safe(prev_file)
        df_today = df_today.loc[:, ~df_today.columns.str.contains('^Unnamed')]
        df_last = df_last.loc[:, ~df_last.columns.str.contains('^Unnamed')]
        mapping = {}
        for _, row in df_last.iterrows():
            cid = str(row.get('合同编号', '')).strip()
            if cid and cid != 'nan': mapping[cid] = {'r': row.get('逾期具体原因', ''), 'c': row.get('逾期原因分类', '')}
        df_today['合同编号'] = df_today['合同编号'].astype(str).str.strip()
        df_today["逾期具体原因_新"] = df_today["合同编号"].apply(lambda x: mapping.get(x, {}).get('r', ''))
        df_today["逾期原因分类_新"] = df_today["合同编号"].apply(lambda x: mapping.get(x, {}).get('c', ''))
        mask_empty = df_today["逾期原因分类_新"] == ""
        if mask_empty.any():
            clause_col = "是否约定保证金条款"
            if clause_col in df_today.columns:
                df_today.loc[mask_empty & (df_today[clause_col] == "是"), ["逾期具体原因_新", "逾期原因分类_新"]] = ["保证金待收取，已催收", "A实际已逾期：指未按合同约定及时足额支付初始保证金。"]
                df_today.loc[mask_empty & (df_today[clause_col] == "否"), ["逾期具体原因_新", "逾期原因分类_新"]] = ["合同未约定收取保证金", "C无需收取保证金：指政策性业务、对养殖户销售业务、分合同、公司批准免收保证金客户的。此类要写明不收取保证金的具体原因。"]
        temp_stream.seek(0)
        book = openpyxl.load_workbook(temp_stream)
        for s in ["WSBZJQKB_Processed", "A类逾期明细", "A类逾期明细汇总"]:
            if s in book.sheetnames: del book[s]
        ws_proc = book.create_sheet("WSBZJQKB_Processed")
        for r in dataframe_to_rows(df_today, index=False, header=True): ws_proc.append(r)
        df_A = df_today[df_today["逾期原因分类_新"] == "A实际已逾期：指未按合同约定及时足额支付初始保证金。"].copy()
        ws_A = book.create_sheet("A类逾期明细")
        for r in dataframe_to_rows(df_A, index=False, header=True): ws_A.append(r)
        clean_and_organize_A_sheet(ws_A)
        optimize_A_sheet_formatting(ws_A)
        today_str = datetime.now().strftime("%Y.%m.%d")
        success, logs = create_A_summary_sheet(book, ws_A, today_str)
        if "WSBZJQKB" in book.sheetnames: fill_original_sheet_columns(book["WSBZJQKB"], df_today)
        if "WSBZJQKB_Processed" in book.sheetnames: del book["WSBZJQKB_Processed"]
        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        return output, logs
    except Exception as e:
        import traceback
        return None, [f"❌ 处理出错: {str(e)}", traceback.format_exc()]
