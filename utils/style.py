import pandas as pd
import io
import copy
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============================================================================
# PART 2: 追加保证金处理逻辑 (ZhuiJIA.py 集成版) - 独立封装，命名加后缀_zj
# ============================================================================

def smart_format_money_zj(value):
    try:
        if pd.isna(value) or value is None: return "0"
        val_float = float(value)
        if abs(val_float) < 0.000001: return "0"
        val_round = round(val_float)
        if val_round == 0: return f"{val_float:.2f}"
        else: return str(val_round)
    except: return str(value)

def smart_format_volume_zj(value, unit="万吨"):
    try:
        if pd.isna(value) or value is None: return f"0{unit}"
        val_float = float(value)
        if 0.0001 <= abs(val_float) < 0.005:
            val_tons = round(val_float * 10000)
            return f"{val_tons}吨"
        s = f"{val_float:.2f}"
        s = s.rstrip('0').rstrip('.')
        if not s: s = "0"
        return f"{s}{unit}"
    except: return f"{str(value)}{unit}"

def smart_format_date_zj(date_obj):
    try:
        if pd.isna(date_obj): return ""
        if isinstance(date_obj, str):
            try: date_obj = pd.to_datetime(date_obj)
            except: return date_obj
        return f"{date_obj.month}月{date_obj.day}日"
    except: return str(date_obj)

def format_number_with_thousands_zj(value):
    try:
        if pd.isna(value) or value is None: return "0"
        num_value = float(value)
        if num_value == 0: return "0"
        int_value = round(num_value)
        if int_value == 0 and abs(num_value) > 0: return f"{num_value:.2f}"
        return f"{int_value:,}"
    except Exception as e: return str(value)

def find_header_row_zj(ws, max_rows_to_check=30):
    key_columns = ['序号', '大区', '经营部', '品种', '客户名称', '合同编号', '合同数量', '合同单价', '调整后待追加保证金金额', '逾期天数', '调整后待执行数量']
    header_candidates = []
    for row_idx in range(1, min(max_rows_to_check, ws.max_row) + 1):
        row_values = []
        key_count = 0
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).strip()
                row_values.append(cell_str)
                for key in key_columns:
                    if key in cell_str:
                        key_count += 1
                        break
        if key_count >= 2 or len(row_values) > 5:
            non_null_count = sum(1 for v in row_values if v and str(v).strip())
            header_candidates.append({'row': row_idx, 'key_count': key_count, 'non_null_count': non_null_count})
    if not header_candidates: return 5
    header_candidates.sort(key=lambda x: (x['key_count'], x['non_null_count']), reverse=True)
    return header_candidates[0]['row']

def get_column_mapping_zj(ws, header_row):
    column_mapping = {}
    reverse_mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        col_name = cell.value
        if not col_name or str(col_name).strip() == '': col_name = f'Unnamed_{col_idx}'
        else: col_name = str(col_name).strip()
        column_mapping[col_idx] = col_name
        reverse_mapping[col_name] = col_idx
    return column_mapping, reverse_mapping

def copy_cell_style_zj(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)
    return target_cell

def apply_excel_like_filtering_zj(ws_original, ws_processed):
    try:
        header_row_idx = find_header_row_zj(ws_original)
        column_mapping, _ = get_column_mapping_zj(ws_original, header_row_idx)

        for col_idx in range(1, ws_original.max_column + 1):
            source = ws_original.cell(row=header_row_idx, column=col_idx)
            target = ws_processed.cell(row=1, column=col_idx)
            target.value = source.value
            copy_cell_style_zj(source, target)
        
        data_rows = []
        for row_idx in range(header_row_idx + 1, ws_original.max_row + 1):
            row_data = []
            for col_idx in range(1, ws_original.max_column + 1):
                row_data.append(ws_original.cell(row=row_idx, column=col_idx).value)
            data_rows.append((row_idx, row_data))
        
        am_idx = ak_idx = as_idx = b_idx = None
        for idx, name in column_mapping.items():
            s = str(name)
            if '调整后待追加保证金金额' in s: am_idx = idx
            elif '待执行数量调整原因分类' in s: ak_idx = idx
            elif '逾期未回款原因分类' in s: as_idx = idx
            elif '大区' in s and '玉米中心' not in s: b_idx = idx

        if not all([am_idx, ak_idx, as_idx, b_idx]): return [], column_mapping

        filtered_rows = []
        for r_idx, r_data in data_rows:
            include = True
            try:
                val = float(r_data[am_idx-1]) if r_data[am_idx-1] is not None else 0
                if val <= 0.01: include = False
            except: include = False

            if include:
                val = str(r_data[ak_idx-1]) if r_data[ak_idx-1] else ""
                if val in ["合同不再继续执行", "合同约定免收追加保证金"]: include = False
            
            if include:
                val = str(r_data[as_idx-1]) if r_data[as_idx-1] else ""
                if val == "C:追加保证金实际已收到，尚未认领": include = False

            if include:
                val = str(r_data[b_idx-1]) if r_data[b_idx-1] else ""
                if val == "玉米中心": include = False

            if include:
                filtered_rows.append((r_idx, r_data))

        curr_row = 2
        for s_idx, r_data in filtered_rows:
            for c_idx, val in enumerate(r_data, 1):
                s_cell = ws_original.cell(row=s_idx, column=c_idx)
                t_cell = ws_processed.cell(row=curr_row, column=c_idx)
                t_cell.value = val
                copy_cell_style_zj(s_cell, t_cell)
            curr_row += 1

        for col in range(1, ws_original.max_column + 1):
            col_letter = get_column_letter(col)
            if ws_original.column_dimensions[col_letter].width:
                ws_processed.column_dimensions[col_letter].width = ws_original.column_dimensions[col_letter].width
        
        ws_processed.freeze_panes = 'A2'
        return filtered_rows, column_mapping
    except: return [], {}

def generate_analysis_report_zj(df_processed, today_display):
    try:
        d_col = b_col = exec_qty_col = am_col = trigger_date_col = an_col = deposit_type_col = None
        for col_name in df_processed.columns:
            col_str = str(col_name)
            if '细分品种' in col_str: d_col = col_name
            elif '大区' in col_str and '玉米中心' not in col_str: b_col = col_name
            elif '调整后待执行数量' in col_str: exec_qty_col = col_name
            elif '调整后待追加保证金金额' in col_str: am_col = col_name
            elif ('追加保证金触发日期' in col_str or '触发日期' in col_str) and '到期' not in col_str: trigger_date_col = col_name
            elif '逾期' in col_str and '天' in col_str: an_col = col_name
            elif '保证金类型' in col_str: deposit_type_col = col_name

        if not exec_qty_col or not am_col: return "分析报告生成失败：缺少必要的列数据。"

        df_processed[exec_qty_col] = pd.to_numeric(df_processed[exec_qty_col], errors='coerce')
        df_processed[am_col] = pd.to_numeric(df_processed[am_col], errors='coerce')
        if an_col: df_processed[an_col] = pd.to_numeric(df_processed[an_col], errors='coerce')
        if trigger_date_col: df_processed[trigger_date_col] = pd.to_datetime(df_processed[trigger_date_col], errors='coerce')

        total_contracts = len(df_processed)
        total_exec_qty = df_processed[exec_qty_col].sum() / 10000
        total_am_amount = df_processed[am_col].sum()

        deposit_amount_parts = []
        if deposit_type_col:
            df_processed[deposit_type_col] = df_processed[deposit_type_col].astype(str)
            down_deposit = df_processed[df_processed[deposit_type_col].str.contains('跌价', na=False)][am_col].sum()
            up_deposit = df_processed[df_processed[deposit_type_col].str.contains('涨价', na=False)][am_col].sum()
            if down_deposit > 0.000001: deposit_amount_parts.append(f"应收取跌价保证金{smart_format_money_zj(down_deposit)}万元")
            if up_deposit > 0.000001: deposit_amount_parts.append(f"应收取涨价保证金{smart_format_money_zj(up_deposit)}万元")
            if not deposit_amount_parts: deposit_amount_str = f"应收取追加保证金{smart_format_money_zj(total_am_amount)}万元"
            else: deposit_amount_str = "、".join(deposit_amount_parts)
        else: deposit_amount_str = f"应收取追加保证金{smart_format_money_zj(total_am_amount)}万元"

        product_summary = []
        if d_col:
            for product, group in df_processed.groupby(d_col):
                amt = group[am_col].sum()
                if amt > 0.000001: product_summary.append(f"{product}{smart_format_money_zj(amt)}万元")
        product_summary_str = "，".join(product_summary)

        trigger_date_summary = []
        trigger_date_summary_str = ""
        overdue_contracts = 0
        overdue_amount = 0
        if trigger_date_col:
            df_sorted = df_processed.sort_values(by=trigger_date_col)
            if an_col:
                mask = df_processed[an_col] > 0
                overdue_contracts = mask.sum()
                if overdue_contracts > 0: overdue_amount = df_processed.loc[mask, am_col].sum()
            for date, group in df_sorted.groupby(trigger_date_col):
                if pd.notnull(date):
                    date_str = smart_format_date_zj(date)
                    d_amt = group[am_col].sum()
                    o_str = ""
                    if an_col:
                        od = group[an_col].dropna()
                        if not od.empty and od.max() > 0: o_str = f"（逾期{int(od.max())}天）"
                    if d_amt > 0.000001: trigger_date_summary.append(f"{date_str}触发{smart_format_money_zj(d_amt)}万元{o_str}")
            trigger_date_summary_str = "，".join(trigger_date_summary)

        region_summary = []
        if b_col:
            r_data = []
            for region, group in df_processed.groupby(b_col):
                r_data.append({'region': region, 'contracts': len(group), 'exec_qty': group[exec_qty_col].sum()/10000, 'am_amount': group[am_col].sum()})
            r_data.sort(key=lambda x: x['am_amount'], reverse=True)
            for i, r in enumerate(r_data, 1):
                region_summary.append(f"{i}、{r['region']}：{r['contracts']}笔，待执行数量{smart_format_volume_zj(r['exec_qty'])}，需追加保证金金额{smart_format_money_zj(r['am_amount'])}万元。")
        region_summary_str = "\n".join(region_summary)

        report_base = f"""截至{today_display}，存续追加保证金合同{total_contracts}笔，对应待执行量{smart_format_volume_zj(total_exec_qty)}，{deposit_amount_str}"""
        if product_summary_str: report_base += f"。分品种看，{product_summary_str}"
        if overdue_contracts > 0: report_base += f"。其中，{overdue_contracts}笔合同已逾期，逾期金额{smart_format_money_zj(overdue_amount)}万元"
        if trigger_date_summary_str:
            sep = "。" if overdue_contracts > 0 else "。其中，"
            report_base += f"{sep}{trigger_date_summary_str}"
        return report_base + f"。分大区情况如下：\n{region_summary_str}"
    except: return "分析报告生成失败。"

def generate_customer_analysis_report_zj(df_processed, today_display):
    try:
        c_col = b_col = exec_qty_col = am_col = an_col = deposit_type_col = None
        for col_name in df_processed.columns:
            col_str = str(col_name)
            if '客户' in col_str and '名称' in col_str: c_col = col_name
            elif '大区' in col_str and '玉米中心' not in col_str: b_col = col_name
            elif '调整后待执行数量' in col_str: exec_qty_col = col_name
            elif '调整后待追加保证金金额' in col_str: am_col = col_name
            elif '逾期' in col_str and '天' in col_str: an_col = col_name
            elif '保证金类型' in col_str: deposit_type_col = col_name

        if not c_col or not am_col: return "客户分析报告生成失败：缺少必要的列数据。"

        df_processed[exec_qty_col] = pd.to_numeric(df_processed[exec_qty_col], errors='coerce')
        df_processed[am_col] = pd.to_numeric(df_processed[am_col], errors='coerce')
        if an_col: df_processed[an_col] = pd.to_numeric(df_processed[an_col], errors='coerce')

        total_am_fmt = format_number_with_thousands_zj(df_processed[am_col].sum())
        report_header = f"截至{today_display}，存续追加保证金合同{len(df_processed)}笔，待执行数量{smart_format_volume_zj(df_processed[exec_qty_col].sum()/10000)}，需追加保证金金额{total_am_fmt}万元。"

        c_data = []
        for customer, group in df_processed.groupby(c_col):
            if pd.isna(customer) or customer == "": continue
            regions_str = "、".join([str(r) for r in (group[b_col].dropna().unique() if b_col else []) if pd.notna(r)])
            d_types = ""
            if deposit_type_col:
                dt = group[deposit_type_col].dropna().unique()
                dt_str = "、".join([str(d) for d in dt if pd.notna(d) and str(d).strip() != ""])
                if dt_str: d_types = f"{dt_str}，"
            max_od = group[an_col].max() if an_col else 0
            if pd.isna(max_od): max_od = 0
            
            c_data.append({
                'customer': customer, 'regions': regions_str, 'contracts': len(group),
                'exec_qty': group[exec_qty_col].sum()/10000, 'am_amount': group[am_col].sum(),
                'max_overdue': max_od, 'am_fmt': format_number_with_thousands_zj(group[am_col].sum()),
                'max_od_str': str(round(max_od)), 'd_types': d_types
            })

        c_data.sort(key=lambda x: (-x['max_overdue'], -x['contracts'], -x['exec_qty']))
        c_summary = []
        for i, info in enumerate(c_data, 1):
            od_s = f"，最长逾期{info['max_od_str']}天" if info['max_overdue'] > 0 else ""
            # 修改点：在 line 的末尾添加双换行符 \n\n
            line = f"{i}、{info['regions']}：{info['contracts']}笔，{info['customer']}，{info['d_types']}待执行数量{smart_format_volume_zj(info['exec_qty'])}，需追加保证金金额{info['am_fmt']}万元{od_s}。\n\n"
            c_summary.append(line)
        return f"{report_header}\n\n{'\n'.join(c_summary)}"
    except: return "客户分析报告生成失败。"

def generate_region_department_report_zj(df_region, today_display, region_name):
    try:
        exec_qty_col = am_col = d_col = trigger_date_col = an_col = dept_col = deposit_type_col = None
        for col_name in df_region.columns:
            col_str = str(col_name)
            if '调整后待执行数量' in col_str: exec_qty_col = col_name
            elif '调整后待追加保证金金额' in col_str: am_col = col_name
            elif '细分品种' in col_str: d_col = col_name
            elif ('追加保证金触发日期' in col_str or '触发日期' in col_str) and '到期' not in col_str: trigger_date_col = col_name
            elif '逾期' in col_str and '天' in col_str: an_col = col_name
            elif '经营部' in col_str: dept_col = col_name
            elif '保证金类型' in col_str: deposit_type_col = col_name

        if not exec_qty_col or not am_col: return f"{region_name}大区报告生成失败：缺少必要列数据。"

        df_region[exec_qty_col] = pd.to_numeric(df_region[exec_qty_col], errors='coerce')
        df_region[am_col] = pd.to_numeric(df_region[am_col], errors='coerce')
        if an_col: df_region[an_col] = pd.to_numeric(df_region[an_col], errors='coerce')
        if trigger_date_col: df_region[trigger_date_col] = pd.to_datetime(df_region[trigger_date_col], errors='coerce')

        total_exec_qty = df_region[exec_qty_col].sum() / 10000
        total_am_amount = df_region[am_col].sum()

        deposit_amount_parts = []
        if deposit_type_col:
            df_region[deposit_type_col] = df_region[deposit_type_col].astype(str)
            down_deposit = df_region[df_region[deposit_type_col].str.contains('跌价', na=False)][am_col].sum()
            up_deposit = df_region[df_region[deposit_type_col].str.contains('涨价', na=False)][am_col].sum()
            if down_deposit > 0.000001: deposit_amount_parts.append(f"应收取跌价保证金{smart_format_money_zj(down_deposit)}万元")
            if up_deposit > 0.000001: deposit_amount_parts.append(f"应收取涨价保证金{smart_format_money_zj(up_deposit)}万元")
            if not deposit_amount_parts: deposit_amount_str = f"应收取追加保证金{smart_format_money_zj(total_am_amount)}万元"
            else: deposit_amount_str = "、".join(deposit_amount_parts)
        else: deposit_amount_str = f"应收取追加保证金{smart_format_money_zj(total_am_amount)}万元"

        prod_summary_str = ""
        if d_col:
            prods = []
            for p, g in df_region.groupby(d_col):
                amt = g[am_col].sum()
                if amt > 0.000001: prods.append(f"{p}{smart_format_money_zj(amt)}万元")
            prod_summary_str = "，".join(prods)

        trigger_str = ""
        if trigger_date_col:
            t_sums = []
            df_sorted = df_region.sort_values(by=trigger_date_col)
            for date, group in df_sorted.groupby(trigger_date_col):
                if pd.notnull(date):
                    d_amt = group[am_col].sum()
                    o_str = ""
                    if an_col:
                        od = group[an_col].dropna()
                        if not od.empty and od.max() > 0: o_str = f"（逾期{int(od.max())}天）"
                    if d_amt > 0.000001: t_sums.append(f"{smart_format_date_zj(date)}触发{smart_format_money_zj(d_amt)}万元{o_str}")
            trigger_str = "，".join(t_sums)

        overdue_contracts = 0
        overdue_amount = 0
        if an_col:
            mask = df_region[an_col] > 0
            overdue_contracts = mask.sum()
            if overdue_contracts > 0: overdue_amount = df_region.loc[mask, am_col].sum()

        dept_str = ""
        if dept_col:
            d_data = []
            for dept, group in df_region.groupby(dept_col):
                d_data.append({'dept': dept, 'contracts': len(group), 'exec_qty': group[exec_qty_col].sum()/10000, 'am_amount': group[am_col].sum()})
            d_data.sort(key=lambda x: x['am_amount'], reverse=True)
            d_lines = []
            for i, d in enumerate(d_data, 1):
                name = d['dept'] if pd.notna(d['dept']) and d['dept'] != "" else "未知经营部"
                d_lines.append(f"{i}、{name}：{d['contracts']}笔，待执行数量{smart_format_volume_zj(d['exec_qty'])}，需追加保证金金额{smart_format_money_zj(d['am_amount'])}万元。")
            dept_str = "\n".join(d_lines)

        report_base = f"""截至{today_display}，{region_name}存续追加保证金合同{len(df_region)}笔，对应待执行量{smart_format_volume_zj(total_exec_qty)}，{deposit_amount_str}"""
        if prod_summary_str: report_base += f"。分品种看，{prod_summary_str}"
        if overdue_contracts > 0: report_base += f"。其中，{overdue_contracts}笔合同已逾期，逾期金额{smart_format_money_zj(overdue_amount)}万元"
        if trigger_str:
            sep = "。" if overdue_contracts > 0 else "。其中，"
            report_base += f"{sep}{trigger_str}"
        return report_base + f"。分经营部情况如下：\n{dept_str}"
    except: return f"{region_name}大区报告生成失败。"

def generate_region_customer_report_zj(df_region, today_display, region_name):
    try:
        c_col = exec_qty_col = am_col = an_col = dept_col = deposit_type_col = None
        for col_name in df_region.columns:
            col_str = str(col_name)
            if '客户' in col_str and '名称' in col_str: c_col = col_name
            elif '调整后待执行数量' in col_str: exec_qty_col = col_name
            elif '调整后待追加保证金金额' in col_str: am_col = col_name
            elif '逾期' in col_str and '天' in col_str: an_col = col_name
            elif '经营部' in col_str: dept_col = col_name
            elif '保证金类型' in col_str: deposit_type_col = col_name

        if not c_col or not am_col: return f"{region_name}大区客户分析报告生成失败。"

        df_region[exec_qty_col] = pd.to_numeric(df_region[exec_qty_col], errors='coerce')
        df_region[am_col] = pd.to_numeric(df_region[am_col], errors='coerce')
        if an_col: df_region[an_col] = pd.to_numeric(df_region[an_col], errors='coerce')

        total_am_fmt = format_number_with_thousands_zj(df_region[am_col].sum())
        report_header = f"截至{today_display}，{region_name}存续追加保证金合同{len(df_region)}笔，待执行数量{smart_format_volume_zj(df_region[exec_qty_col].sum()/10000)}，需追加保证金金额{total_am_fmt}万元。"

        c_data = []
        for customer, group in df_region.groupby(c_col):
            if pd.isna(customer) or customer == "": continue
            depts_str = ""
            if dept_col:
                depts = group[dept_col].dropna().unique()
                depts_str = "、".join([str(d) for d in depts if pd.notna(d) and str(d).strip() != ""])
            d_types_str = ""
            if deposit_type_col:
                dt = group[deposit_type_col].dropna().unique()
                t_str = "、".join([str(t) for t in dt if pd.notna(t) and str(t).strip() != ""])
                if t_str: d_types_str = f"{t_str}，"
            max_od = group[an_col].max() if an_col else 0
            if pd.isna(max_od): max_od = 0
            c_data.append({
                'customer': customer, 'depts': depts_str, 'contracts': len(group),
                'exec_qty': group[exec_qty_col].sum()/10000, 'am_amount': group[am_col].sum(),
                'max_overdue': max_od, 'am_fmt': format_number_with_thousands_zj(group[am_col].sum()),
                'max_od_str': str(round(max_od)), 'd_types': d_types_str
            })

        c_data.sort(key=lambda x: (-x['max_overdue'], -x['contracts'], -x['exec_qty']))
        lines = []
        for i, info in enumerate(c_data, 1):
            od_s = f"，最长逾期{info['max_od_str']}天" if info['max_overdue'] > 0 else ""
            prefix = f"{i}、{info['depts']}：" if info['depts'] else f"{i}、"
            # 修改点：在 line 的末尾添加双换行符 \n\n
            lines.append(f"{prefix}{info['contracts']}笔，{info['customer']}，{info['d_types']}待执行数量{smart_format_volume_zj(info['exec_qty'])}，需追加保证金金额{info['am_fmt']}万元{od_s}。\n\n")
        return f"{report_header}\n\n{'\n'.join(lines)}"
    except: return f"{region_name}大区客户分析报告生成失败。"

def process_additional_margin_logic(uploaded_file, region_filter):
    logs = []
    try:
        today_display = f"{datetime.now().month}月{datetime.now().day}日"
        book = openpyxl.load_workbook(uploaded_file)
        ws_original = book.worksheets[0] 
        if '追保处理' in book.sheetnames: del book['追保处理']
        ws_processed = book.create_sheet('追保处理')
        filtered_rows, column_names = apply_excel_like_filtering_zj(ws_original, ws_processed)
        if not filtered_rows:
            return None, ["⚠️ 警告：筛选后没有数据行！"], "", ""
        data_for_analysis = []
        for _, row_data in filtered_rows:
            row_dict = {}
            for col_idx, value in enumerate(row_data, 1):
                if col_idx in column_names:
                    row_dict[column_names[col_idx]] = value
            data_for_analysis.append(row_dict)
        df_processed = pd.DataFrame(data_for_analysis)
        if '分析报告' in book.sheetnames: del book['分析报告']
        ws_report = book.create_sheet('分析报告')
        b_col = next((c for c in df_processed.columns if '大区' in str(c) and '玉米中心' not in str(c)), None)
        report_A = ""
        report_B = ""
        if region_filter == "中粮贸易":
            report_A = generate_analysis_report_zj(df_processed, today_display)
            report_B = generate_customer_analysis_report_zj(df_processed, today_display)
        else:
            if not b_col:
                return None, [f"❌ 数据中找不到“大区”列，无法进行大区筛选。"], "", ""
            df_region = df_processed[df_processed[b_col] == region_filter].copy()
            if len(df_region) == 0:
                return None, [f"⚠️ 筛选结果中没有包含【{region_filter}】的数据。"], "", ""
            report_A = generate_region_department_report_zj(df_region, today_display, region_filter)
            report_B = generate_region_customer_report_zj(df_region, today_display, region_filter)
        ws_report.cell(row=1, column=1, value=report_A)
        ws_report.cell(row=1, column=2, value=report_B)
        ws_report.column_dimensions['A'].width = 100
        ws_report.column_dimensions['B'].width = 100
        for row in ws_report.iter_rows():
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.font = Font(size=10, name='宋体')
                    ws_report.row_dimensions[cell.row].height = 200
        ws_report.freeze_panes = 'A2'
        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        logs.append(f"✅ 【{region_filter}】分析报告生成成功！")
        return output, logs, report_A, report_B
    except Exception as e:
        import traceback
        return None, [f"❌ 处理出错: {str(e)}", traceback.format_exc()], "", ""
